from barrier_func_method import barrier_method
from main import f, g

import openpyxl
from decimal import *

getcontext().prec = 4


def f_range(start: Decimal, stop: Decimal, step: Decimal):
    while start <= stop:
        yield Decimal(start)
        start += step


def test(sheet):
    x0 = [Decimal(0.1), Decimal(0.5), Decimal(0.1), Decimal(0.2), Decimal(0.2)]
    mu0 = Decimal(10)
    beta = Decimal(0.4)

    row = 1
    sheet.cell(row=row, column=1).value = 'set of starting points'
    row += 1
    for i, xi in enumerate(x0):

        print(f'x{i + 1} variable is fixed')

        sheet.cell(row=row, column=1).value = f'x{i + 1} variable is fixed'
        sheet.cell(row=row, column=2).value = 'x_min'
        sheet.cell(row=row, column=3).value = 'f(x_min)'
        sheet.cell(row=row, column=4).value = 'Number of iterations'
        row += 1
        for j in f_range(Decimal(-10), Decimal(10), Decimal(0.1)):
            x0[i] = j
            if g.check_point(x0):
                xk, k = barrier_method(f, g, x0, mu0, beta, barrier_type='logarithmic', points='valid')
                if xk != [0, 0, 0, 0, 0]:
                    sheet.cell(row=row,
                               column=1).value = f'x0 = ({x0[0]:.1f}, {x0[1]:.1f}, {x0[2]:.1f}, {x0[3]:.1f}, {x0[4]:.1f})'
                    sheet.cell(row=row,
                               column=2).value = f'({xk[0]:.1f}, {xk[1]:.1f}, {xk[2]:.1f}, {xk[3]:.1f}, {xk[4]:.1f})'
                    sheet.cell(row=row, column=3).value = f.get_func()(xk)
                    sheet.cell(row=row, column=4).value = k
                    row += 1
            x0[i] = xi


def test2(sheet2):
    x0 = [Decimal(0), Decimal(0), Decimal(0), Decimal(0), Decimal(0)]
    mu0 = Decimal(10)
    beta = Decimal(0.4)

    row = 1
    sheet2.cell(row=row, column=1).value = 'x0'
    sheet2.cell(row=row, column=2).value = 'x_min'
    sheet2.cell(row=row, column=3).value = 'f(x_min)'
    sheet2.cell(row=row, column=4).value = 'Number of iterations'
    row += 1
    for value in f_range(Decimal(-50), Decimal(50), Decimal(1)):
        x0[0] = value
        x0[1] = value
        x0[2] = value
        x0[3] = value
        x0[4] = value
        if x0 != [Decimal(0), Decimal(0), Decimal(0), Decimal(0), Decimal(0)]:
            xk, k = barrier_method(f, g, x0, mu0, beta, barrier_type='logarithmic', points='all')
            sheet2.cell(row=row,
                        column=1).value = f'({x0[0]:.1f}, {x0[1]:.1f}, {x0[2]:.1f}, {x0[3]:.1f}, {x0[4]:.1f})'
            sheet2.cell(row=row,
                        column=2).value = f'({xk[0]:.1f}, {xk[1]:.1f}, {xk[2]:.1f}, {xk[3]:.1f}, {xk[4]:.1f})'
            sheet2.cell(row=row, column=3).value = f.get_func()(xk)
            sheet2.cell(row=row, column=4).value = k
            row += 1


def main():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    test(ws1)
    # ws2 = wb.create_sheet()
    # test2(ws2)
    wb.save(filename="excel.xlsx")

    return 0


if __name__ == '__main__':
    main()
